"""
Turbo.az Analytics - Export Module
CSV ve Excel export fonksiyonlari
"""

import csv
import io
from datetime import datetime
from typing import Dict, List, Optional

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


def format_car_for_export(car: Dict) -> Dict:
    """Arac verisini export icin formatla."""
    return {
        "ID": car.get("turbo_id", ""),
        "Arac Adi": car.get("name", ""),
        "Marka": car.get("brand", ""),
        "Model": car.get("model", ""),
        "Fiyat": car.get("price", 0),
        "Para Birimi": car.get("currency", "AZN"),
        "Yil": car.get("year", ""),
        "Motor": car.get("engine", ""),
        "Kilometre": car.get("mileage", ""),
        "Sehir": car.get("city", ""),
        "Goruntulenme": car.get("views", 0),
        "Sifir KM": "Evet" if car.get("is_new") else "Hayir",
        "VIP": "Evet" if car.get("is_vip") else "Hayir",
        "Premium": "Evet" if car.get("is_premium") else "Hayir",
        "URL": car.get("url", ""),
        "Eklenme Tarihi": car.get("created_at", ""),
        "Guncelleme Tarihi": car.get("updated_at", ""),
    }


def export_cars_csv(cars: List[Dict]) -> bytes:
    """Araclari CSV formatinda export et."""
    if not cars:
        return b""

    output = io.StringIO()
    formatted_cars = [format_car_for_export(car) for car in cars]

    fieldnames = list(formatted_cars[0].keys())
    writer = csv.DictWriter(output, fieldnames=fieldnames)

    writer.writeheader()
    writer.writerows(formatted_cars)

    return output.getvalue().encode("utf-8-sig")  # BOM ile UTF-8


def export_cars_excel(cars: List[Dict]) -> Optional[bytes]:
    """Araclari Excel formatinda export et."""
    if not EXCEL_AVAILABLE:
        return None

    if not cars:
        return b""

    wb = Workbook()
    ws = wb.active
    ws.title = "Araclar"

    formatted_cars = [format_car_for_export(car) for car in cars]
    headers = list(formatted_cars[0].keys())

    # Header stili
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    # Header'lari yaz
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Verileri yaz
    for row, car in enumerate(formatted_cars, 2):
        for col, header in enumerate(headers, 1):
            ws.cell(row=row, column=col, value=car[header])

    # Sutun genisliklerini ayarla
    for col, header in enumerate(headers, 1):
        max_length = len(str(header))
        for row in range(2, min(len(formatted_cars) + 2, 100)):
            cell_value = ws.cell(row=row, column=col).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        ws.column_dimensions[chr(64 + col) if col <= 26 else f"A{chr(64 + col - 26)}"].width = min(max_length + 2, 50)

    # Bytes olarak dondur
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def export_price_history_csv(history: List[Dict]) -> bytes:
    """Fiyat gecmisini CSV formatinda export et."""
    if not history:
        return b""

    output = io.StringIO()

    fieldnames = ["Tarih", "Eski Fiyat", "Yeni Fiyat", "Para Birimi", "Degisim", "Degisim %"]

    writer = csv.DictWriter(output, fieldnames=fieldnames)
    writer.writeheader()

    for item in history:
        old_price = item.get("old_price", 0) or 0
        new_price = item.get("new_price", 0) or 0
        change = new_price - old_price
        change_percent = (change / old_price * 100) if old_price else 0

        writer.writerow({
            "Tarih": item.get("recorded_at", ""),
            "Eski Fiyat": old_price,
            "Yeni Fiyat": new_price,
            "Para Birimi": item.get("currency", "AZN"),
            "Degisim": change,
            "Degisim %": f"{change_percent:.2f}%",
        })

    return output.getvalue().encode("utf-8-sig")


def generate_export_filename(prefix: str, extension: str) -> str:
    """Export dosyasi icin benzersiz isim olustur."""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"{prefix}_{timestamp}.{extension}"
